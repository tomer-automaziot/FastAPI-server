"""
Excel Image Extractor API v2
Extracts embedded images from Excel files (both .xls and .xlsx) and returns them as base64-encoded data.
Handles legacy .xls files by converting to .xlsx first using LibreOffice.
Designed to work with n8n workflows and Priority ERP integration.
"""

from flask import Flask, request, jsonify
from openpyxl import load_workbook
from PIL import Image
import io
import base64
import tempfile
import os
import subprocess
import logging
import shutil

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Max file size: 50MB
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024


def detect_file_format(file_bytes):
    """
    Detect if file is XLS (legacy) or XLSX (modern) format.
    
    Returns:
        'xls' - Legacy Excel format (needs conversion)
        'xlsx' - Modern Excel format (can process directly)
        'unknown' - Unknown format
    """
    # XLSX files are ZIP archives starting with PK
    if file_bytes[:4] == b'PK\x03\x04':
        return 'xlsx'
    
    # XLS files are OLE Compound Documents starting with D0 CF 11 E0
    if file_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return 'xls'
    
    return 'unknown'


def convert_xls_to_xlsx(xls_path, output_dir):
    """
    Convert .xls file to .xlsx using LibreOffice.
    
    Args:
        xls_path: Path to the .xls file
        output_dir: Directory for the output .xlsx file
        
    Returns:
        Path to the converted .xlsx file, or None if conversion failed
    """
    try:
        logger.info(f"Converting XLS to XLSX: {xls_path}")
        
        # Run LibreOffice conversion
        result = subprocess.run([
            'libreoffice',
            '--headless',
            '--convert-to', 'xlsx',
            '--outdir', output_dir,
            xls_path
        ], capture_output=True, text=True, timeout=120)
        
        if result.returncode != 0:
            logger.error(f"LibreOffice conversion failed: {result.stderr}")
            return None
        
        # Find the converted file
        base_name = os.path.splitext(os.path.basename(xls_path))[0]
        xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")
        
        if os.path.exists(xlsx_path):
            logger.info(f"Conversion successful: {xlsx_path}")
            return xlsx_path
        else:
            logger.error(f"Converted file not found: {xlsx_path}")
            return None
            
    except subprocess.TimeoutExpired:
        logger.error("LibreOffice conversion timed out")
        return None
    except Exception as e:
        logger.error(f"Conversion error: {str(e)}")
        return None


def extract_images_from_xlsx(xlsx_path):
    """
    Extract all images from an XLSX file.
    
    Args:
        xlsx_path: Path to the .xlsx file
        
    Returns:
        List of dictionaries containing image data and metadata
    """
    images = []
    
    try:
        wb = load_workbook(xlsx_path)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            if hasattr(sheet, '_images') and sheet._images:
                logger.info(f"Sheet '{sheet_name}' has {len(sheet._images)} images")
                
                for idx, img in enumerate(sheet._images):
                    try:
                        img_data = None
                        
                        # Try different methods to get image data
                        if hasattr(img, '_data') and callable(img._data):
                            img_data = img._data()
                        elif hasattr(img, 'ref'):
                            if hasattr(img.ref, 'getvalue'):
                                img_data = img.ref.getvalue()
                            elif hasattr(img.ref, 'read'):
                                img_data = img.ref.read()
                        
                        if not img_data:
                            logger.warning(f"Could not extract data for image {idx} in sheet {sheet_name}")
                            continue
                        
                        # Detect format
                        if img_data[:8] == b'\x89PNG\r\n\x1a\n':
                            img_format = 'png'
                        elif img_data[:2] == b'\xff\xd8':
                            img_format = 'jpeg'
                        elif img_data[:4] == b'GIF8':
                            img_format = 'gif'
                        else:
                            img_format = 'png'  # Default
                        
                        # Get position info
                        anchor_cell = None
                        if hasattr(img, 'anchor'):
                            if hasattr(img.anchor, '_from'):
                                col = img.anchor._from.col
                                row = img.anchor._from.row
                                anchor_cell = f"{chr(65 + min(col, 25))}{row + 1}"
                        
                        # Get dimensions
                        width = None
                        height = None
                        try:
                            pil_img = Image.open(io.BytesIO(img_data))
                            width, height = pil_img.size
                            
                            # Re-save as PNG to ensure clean data
                            output_buffer = io.BytesIO()
                            pil_img.save(output_buffer, format='PNG')
                            img_data = output_buffer.getvalue()
                            img_format = 'png'
                        except Exception as e:
                            logger.warning(f"Could not process image with PIL: {e}")
                        
                        # Encode to base64
                        b64_data = base64.b64encode(img_data).decode('utf-8')
                        
                        images.append({
                            'index': len(images),
                            'sheet': sheet_name,
                            'cell': anchor_cell,
                            'format': img_format,
                            'mime_type': f'image/{img_format}',
                            'width': width,
                            'height': height,
                            'size_bytes': len(img_data),
                            'base64': b64_data,
                            'data_uri': f'data:image/{img_format};base64,{b64_data}'
                        })
                        
                        logger.info(f"Extracted image {len(images)}: {width}x{height} from sheet '{sheet_name}'")
                    
                    except Exception as e:
                        logger.error(f"Error extracting image {idx} from sheet {sheet_name}: {str(e)}")
                        continue
        
        wb.close()
        
    except Exception as e:
        logger.error(f"Error processing XLSX file: {str(e)}")
        raise
    
    return images


def extract_images_from_excel(file_bytes, filename='unknown'):
    """
    Extract all images from an Excel file (handles both XLS and XLSX).
    
    Args:
        file_bytes: Binary content of the Excel file
        filename: Original filename (for logging)
        
    Returns:
        List of dictionaries containing image data and metadata
    """
    # Detect file format
    file_format = detect_file_format(file_bytes)
    logger.info(f"Detected file format: {file_format} for {filename}")
    
    # Create temp directory for processing
    temp_dir = tempfile.mkdtemp()
    
    try:
        if file_format == 'xls':
            # Save XLS file
            xls_path = os.path.join(temp_dir, 'input.xls')
            with open(xls_path, 'wb') as f:
                f.write(file_bytes)
            
            # Convert to XLSX
            xlsx_path = convert_xls_to_xlsx(xls_path, temp_dir)
            
            if not xlsx_path:
                raise Exception("Failed to convert XLS to XLSX. LibreOffice conversion failed.")
            
            # Extract images from converted XLSX
            images = extract_images_from_xlsx(xlsx_path)
            
        elif file_format == 'xlsx':
            # Save XLSX file directly
            xlsx_path = os.path.join(temp_dir, 'input.xlsx')
            with open(xlsx_path, 'wb') as f:
                f.write(file_bytes)
            
            # Extract images
            images = extract_images_from_xlsx(xlsx_path)
            
        else:
            raise Exception(f"Unknown file format. Expected XLS or XLSX, got: {file_format}")
        
        return images
        
    finally:
        # Clean up temp directory
        shutil.rmtree(temp_dir, ignore_errors=True)


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint for Railway"""
    # Check if LibreOffice is available
    try:
        result = subprocess.run(['libreoffice', '--version'], capture_output=True, text=True, timeout=10)
        lo_version = result.stdout.strip() if result.returncode == 0 else 'not available'
    except:
        lo_version = 'not available'
    
    return jsonify({
        'status': 'healthy',
        'service': 'excel-image-extractor',
        'version': '2.0',
        'libreoffice': lo_version,
        'features': ['xls_support', 'xlsx_support', 'auto_conversion']
    })


@app.route('/extract', methods=['POST'])
def extract_images():
    """
    Extract images from an uploaded Excel file.
    
    Accepts:
        - multipart/form-data with 'file' field
        - application/json with 'base64' field containing base64-encoded Excel file
        
    Returns:
        JSON with extracted images as base64 data
    """
    try:
        file_bytes = None
        filename = 'unknown.xlsx'
        
        # Check for file upload
        if 'file' in request.files:
            file = request.files['file']
            filename = file.filename or filename
            file_bytes = file.read()
            logger.info(f"Received file upload: {filename} ({len(file_bytes)} bytes)")
        
        # Check for base64 in JSON body
        elif request.is_json:
            data = request.get_json()
            if 'base64' in data:
                file_bytes = base64.b64decode(data['base64'])
                filename = data.get('filename', filename)
                logger.info(f"Received base64 data for: {filename} ({len(file_bytes)} bytes)")
            elif 'data_uri' in data:
                # Handle data URI format
                data_uri = data['data_uri']
                if ',' in data_uri:
                    file_bytes = base64.b64decode(data_uri.split(',')[1])
                filename = data.get('filename', filename)
                logger.info(f"Received data URI for: {filename}")
        
        # Check for raw binary
        elif request.content_type and 'application/octet-stream' in request.content_type:
            file_bytes = request.data
            logger.info(f"Received raw binary data ({len(file_bytes)} bytes)")
        
        if not file_bytes:
            return jsonify({
                'success': False,
                'error': 'No file provided. Send file as multipart/form-data, JSON with base64 field, or raw binary.'
            }), 400
        
        # Detect and log file format
        file_format = detect_file_format(file_bytes)
        logger.info(f"File format detected: {file_format}")
        
        # Extract images
        images = extract_images_from_excel(file_bytes, filename)
        
        logger.info(f"Successfully extracted {len(images)} images from {filename}")
        
        return jsonify({
            'success': True,
            'filename': filename,
            'original_format': file_format,
            'image_count': len(images),
            'images': images
        })
    
    except Exception as e:
        logger.error(f"Error processing request: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/extract-simple', methods=['POST'])
def extract_images_simple():
    """
    Simplified endpoint that returns just the image data URIs.
    Useful for direct integration with Priority ERP.
    
    Returns:
        JSON with array of data URIs ready for Priority attachment upload
    """
    try:
        # Call the main extract function
        response = extract_images()
        
        # Handle tuple response (response, status_code)
        if isinstance(response, tuple):
            if response[1] != 200:
                return response
            data = response[0].get_json()
        else:
            data = response.get_json()
        
        if not data.get('success'):
            return jsonify(data), 400
        
        # Return simplified format for Priority
        return jsonify({
            'success': True,
            'count': data['image_count'],
            'original_format': data.get('original_format', 'unknown'),
            'data_uris': [img['data_uri'] for img in data['images']],
            'images': [{
                'index': img['index'],
                'format': img['format'],
                'width': img.get('width'),
                'height': img.get('height'),
                'data_uri': img['data_uri']
            } for img in data['images']]
        })
    
    except Exception as e:
        logger.error(f"Error in extract-simple: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    logger.info(f"Starting Excel Image Extractor v2 on port {port}")
    app.run(host='0.0.0.0', port=port, debug=False)
